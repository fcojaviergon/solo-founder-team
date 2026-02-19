#!/usr/bin/env python3
"""
Genera el Excel de cotización del PDP.
Uso: python generate_pdp_xlsx.py <output_path> <json_data_path>

El JSON debe tener estructura:
{
  "tasks": [
    {
      "id": "M1-T01",
      "modulo": "Setup & Infraestructura",
      "tarea": "Configurar Docker compose",
      "tipo": "Infra",
      "complejidad": "Baja",
      "hh": 8,
      "perfil": "Senior",
      "dependencias": "",
      "riesgo": "Bajo",
      "fase": "MVP",
      "notas": ""
    }
  ]
}
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='1B2A4A')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
SUB_FILL = PatternFill('solid', fgColor='E8EDF3')
SUB_FONT = Font(bold=True, color='1B2A4A', size=10, name='Arial')
DATA_FONT = Font(color='333333', size=10, name='Arial')
TOTAL_FILL = PatternFill('solid', fgColor='F0F4F8')
TOTAL_FONT = Font(bold=True, color='1B2A4A', size=11, name='Arial')
INPUT_FONT = Font(color='0000FF', size=10, name='Arial')
BORDER = Border(
    left=Side(style='thin', color='D0D5DD'),
    right=Side(style='thin', color='D0D5DD'),
    top=Side(style='thin', color='D0D5DD'),
    bottom=Side(style='thin', color='D0D5DD')
)
COMP_COLORS = {
    'Baja': PatternFill('solid', fgColor='D4EDDA'),
    'Media': PatternFill('solid', fgColor='FFF3CD'),
    'Alta': PatternFill('solid', fgColor='FFE0B2'),
    'Muy Alta': PatternFill('solid', fgColor='F8D7DA'),
}
RISK_COLORS = {
    'Bajo': PatternFill('solid', fgColor='D4EDDA'),
    'Medio': PatternFill('solid', fgColor='FFF3CD'),
    'Alto': PatternFill('solid', fgColor='F8D7DA'),
}

def hdr(ws, row, n):
    for c in range(1, n+1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

def sc(cell, font=None):
    cell.font = font or DATA_FONT
    cell.border = BORDER
    cell.alignment = Alignment(vertical='center', wrap_text=True)

def wbs_sheet(wb, tasks):
    ws = wb.active
    ws.title = "WBS Detallado"
    headers = ['ID','Módulo','Tarea','Tipo','Complejidad','HH','Perfil','Dependencias','Riesgo','MVP/Fase','Notas']
    widths = [10,20,40,12,12,10,16,14,10,12,30]
    ws.merge_cells('A1:K1')
    ws['A1'].value = "WBS — Desglose de Trabajo"
    ws['A1'].font = Font(bold=True, color='1B2A4A', size=14, name='Arial')
    for c,(h,w) in enumerate(zip(headers,widths),1):
        ws.cell(row=3, column=c, value=h)
        ws.column_dimensions[get_column_letter(c)].width = w
    hdr(ws, 3, len(headers))
    cur_mod, row = None, 4
    for t in tasks:
        if t.get('modulo') != cur_mod:
            cur_mod = t['modulo']
            ws.merge_cells(f'A{row}:K{row}')
            ws.cell(row=row, column=1, value=cur_mod).font = SUB_FONT
            for c in range(1,12):
                ws.cell(row=row, column=c).fill = SUB_FILL
                ws.cell(row=row, column=c).border = BORDER
            row += 1
        vals = [t.get('id',''),t.get('modulo',''),t.get('tarea',''),t.get('tipo',''),
                t.get('complejidad',''),t.get('hh',0),t.get('perfil',''),
                t.get('dependencias',''),t.get('riesgo',''),t.get('fase','MVP'),t.get('notas','')]
        for c,v in enumerate(vals,1):
            cell = ws.cell(row=row, column=c, value=v)
            sc(cell)
        comp = ws.cell(row=row, column=5)
        if comp.value in COMP_COLORS: comp.fill = COMP_COLORS[comp.value]
        risk = ws.cell(row=row, column=9)
        if risk.value in RISK_COLORS: risk.fill = RISK_COLORS[risk.value]
        row += 1
    lr = row - 1
    ws.cell(row=row+1, column=5, value="TOTAL HH:").font = TOTAL_FONT
    tc = ws.cell(row=row+1, column=6)
    tc.value = f'=SUM(F4:F{lr})'
    tc.font, tc.fill = TOTAL_FONT, TOTAL_FILL
    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:K{lr}'

def summary_module(wb, tasks):
    ws = wb.create_sheet("Resumen por Módulo")
    ws.merge_cells('A1:D1')
    ws['A1'].value = "Resumen por Módulo"
    ws['A1'].font = Font(bold=True, color='1B2A4A', size=14, name='Arial')
    for c,(h,w) in enumerate(zip(['Módulo','Tareas','HH','%'],[30,12,12,12]),1):
        ws.cell(row=3, column=c, value=h)
        ws.column_dimensions[get_column_letter(c)].width = w
    hdr(ws, 3, 4)
    mods = {}
    for t in tasks:
        m = t.get('modulo','?')
        mods.setdefault(m, {'c':0,'h':0})
        mods[m]['c'] += 1
        mods[m]['h'] += t.get('hh',0)
    total = sum(v['h'] for v in mods.values())
    row = 4
    for m,d in mods.items():
        ws.cell(row=row, column=1, value=m).font = DATA_FONT
        ws.cell(row=row, column=2, value=d['c']).font = DATA_FONT
        ws.cell(row=row, column=3, value=d['h']).font = DATA_FONT
        p = ws.cell(row=row, column=4, value=d['h']/total if total else 0)
        p.font, p.number_format = DATA_FONT, '0.0%'
        for c in range(1,5): ws.cell(row=row, column=c).border = BORDER
        row += 1
    for c in range(1,5):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = BORDER
    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=3, value=total).font = TOTAL_FONT

def summary_profile(wb, tasks):
    ws = wb.create_sheet("Resumen por Perfil")
    ws.merge_cells('A1:D1')
    ws['A1'].value = "Resumen por Perfil"
    ws['A1'].font = Font(bold=True, color='1B2A4A', size=14, name='Arial')
    for c,(h,w) in enumerate(zip(['Perfil','Tareas','HH','%'],[25,12,12,12]),1):
        ws.cell(row=3, column=c, value=h)
        ws.column_dimensions[get_column_letter(c)].width = w
    hdr(ws, 3, 4)
    profs = {}
    for t in tasks:
        p = t.get('perfil','?')
        profs.setdefault(p, {'c':0,'h':0})
        profs[p]['c'] += 1
        profs[p]['h'] += t.get('hh',0)
    total = sum(v['h'] for v in profs.values())
    row = 4
    for p,d in profs.items():
        ws.cell(row=row, column=1, value=p).font = DATA_FONT
        ws.cell(row=row, column=2, value=d['c']).font = DATA_FONT
        ws.cell(row=row, column=3, value=d['h']).font = DATA_FONT
        pc = ws.cell(row=row, column=4, value=d['h']/total if total else 0)
        pc.font, pc.number_format = DATA_FONT, '0.0%'
        for c in range(1,5): ws.cell(row=row, column=c).border = BORDER
        row += 1
    for c in range(1,5):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = BORDER
    ws.cell(row=row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=3, value=total).font = TOTAL_FONT

def cost_calc(wb, tasks):
    ws = wb.create_sheet("Calculadora de Costo")
    ws.merge_cells('A1:F1')
    ws['A1'].value = "Calculadora de Costo"
    ws['A1'].font = Font(bold=True, color='1B2A4A', size=14, name='Arial')
    ws.merge_cells('A2:F2')
    ws['A2'].value = "📝 Ingresa tarifas en columnas C y E (azul = editable)"
    ws['A2'].font = Font(italic=True, color='666666', size=10, name='Arial')
    for c,(h,w) in enumerate(zip(
        ['Perfil','HH','Tarifa/Hr USD','Costo USD','Tarifa/Hr CLP','Costo CLP'],
        [25,12,18,18,18,18]),1):
        ws.cell(row=3, column=c, value=h)
        ws.column_dimensions[get_column_letter(c)].width = w
    hdr(ws, 3, 6)
    profs = {}
    for t in tasks:
        p = t.get('perfil','?')
        profs[p] = profs.get(p, 0) + t.get('hh',0)
    row = 4
    for p,h in profs.items():
        ws.cell(row=row, column=1, value=p).font = DATA_FONT
        ws.cell(row=row, column=2, value=h).font = DATA_FONT
        ws.cell(row=row, column=3, value=0).font = INPUT_FONT
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=f'=B{row}*C{row}').font = DATA_FONT
        ws.cell(row=row, column=4).number_format = '$#,##0'
        ws.cell(row=row, column=5, value=0).font = INPUT_FONT
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=6, value=f'=B{row}*E{row}').font = DATA_FONT
        ws.cell(row=row, column=6).number_format = '$#,##0'
        for c in range(1,7): ws.cell(row=row, column=c).border = BORDER
        row += 1
    lr = row - 1
    for c in range(1,7):
        ws.cell(row=row, column=c).fill = TOTAL_FILL
        ws.cell(row=row, column=c).border = BORDER
    ws.cell(row=row, column=1, value="SUBTOTAL").font = TOTAL_FONT
    ws.cell(row=row, column=2, value=f'=SUM(B4:B{lr})').font = TOTAL_FONT
    ws.cell(row=row, column=4, value=f'=SUM(D4:D{lr})').font = TOTAL_FONT
    ws.cell(row=row, column=4).number_format = '$#,##0'
    ws.cell(row=row, column=6, value=f'=SUM(F4:F{lr})').font = TOTAL_FONT
    ws.cell(row=row, column=6).number_format = '$#,##0'
    sr = row
    row += 2
    ws.cell(row=row, column=1, value="CONTINGENCIA").font = SUB_FONT
    row += 1
    ws.cell(row=row, column=1, value="% Buffer").font = DATA_FONT
    ws.cell(row=row, column=2, value=0.25).font = INPUT_FONT
    ws.cell(row=row, column=2).number_format = '0%'
    br = row
    row += 1
    ws.cell(row=row, column=1, value="HH Buffer").font = DATA_FONT
    ws.cell(row=row, column=2, value=f'=B{sr}*B{br}').font = DATA_FONT
    bhr = row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL CON BUFFER").font = TOTAL_FONT
    ws.cell(row=row, column=1).fill = TOTAL_FILL
    ws.cell(row=row, column=2, value=f'=B{sr}+B{bhr}').font = TOTAL_FONT
    ws.cell(row=row, column=2).fill = TOTAL_FILL
    ws.cell(row=row, column=4, value=f'=D{sr}*(1+B{br})').font = TOTAL_FONT
    ws.cell(row=row, column=4).fill = TOTAL_FILL
    ws.cell(row=row, column=4).number_format = '$#,##0'
    ws.cell(row=row, column=6, value=f'=F{sr}*(1+B{br})').font = TOTAL_FONT
    ws.cell(row=row, column=6).fill = TOTAL_FILL
    ws.cell(row=row, column=6).number_format = '$#,##0'

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Uso: python generate_pdp_xlsx.py <output.xlsx> <data.json>")
        sys.exit(1)
    with open(sys.argv[2], 'r', encoding='utf-8') as f:
        data = json.load(f)
    tasks = data.get('tasks', [])
    wb = Workbook()
    wbs_sheet(wb, tasks)
    summary_module(wb, tasks)
    summary_profile(wb, tasks)
    cost_calc(wb, tasks)
    wb.save(sys.argv[1])
    print(f"✅ Excel generado: {sys.argv[1]}")
